# openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font
import time

book = Workbook()
sheet = book.active

sheet['A1'] = 5
sheet['A2'] = 10

sheet['B1'] = 'rango'
sheet['B1'].font = Font(color='FF0000', bold=True)
for i in range(2, 15):
    sheet[f'B{i}'] = i**2

sheet2 = book.create_sheet('hoja_2')
sheet2['A1'] = 'SUSCRIBETE'
fecha = time.strftime('%x')
sheet2['A2'] = fecha

sheet3 = book.create_sheet('hoja_3')
sheet3.merge_cells('A1:D1')
sheet3['A1'] = 'prueba de union de celdas'
# sheet3.unmerge_cells('A1:D1')


book.save('prueba_escritura.xlsx')